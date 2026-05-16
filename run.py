from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def load_env(path='.env'):
    config = {}
    try:
        with open(path, 'r', encoding='utf-8') as file:
            for raw_line in file:
                line = raw_line.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue
                key, value = line.split('=', 1)
                config[key.strip()] = value.strip().strip('"').strip("'")
    except FileNotFoundError:
        pass
    return config


CONFIG = load_env()


def env(key, default=''):
    return CONFIG.get(key, default)


def env_list(key, default='', sep='|'):
    value = env(key, default)
    return [item.strip() for item in value.split(sep) if item.strip()]


def env_rows(key, default=''):
    rows = []
    for row in env(key, default).split(';'):
        parts = [part.strip() for part in row.split('|')]
        if any(parts):
            rows.append(parts)
    return rows


def color(name, default):
    return env(name, default).replace('#', '').upper()


OUTPUT = env('OUTPUT_FILE', 'Mau_bai_tap_sau_dao_tao_review_luong.xlsx')
MAIN_SHEET = env('MAIN_SHEET', 'Tu danh gia')
REFERENCE_SHEET = env('REFERENCE_SHEET', 'Thang diem va checklist')
REPORT_TITLE = env('REPORT_TITLE')
REPORT_SUBTITLE = env('REPORT_SUBTITLE')
GUIDE_TEXT = env('GUIDE_TEXT')
REFERENCE_TITLE = env('REFERENCE_TITLE')
SECTION_PLACEHOLDER = env('SECTION_PLACEHOLDER')
CHECKLIST_MARK = env('CHECKLIST_MARK', '☐')
RATING_FALLBACK = env('RATING_FALLBACK', 'Theo dải điểm tài liệu')
TOTAL_SELF_SCORE_LABEL = env('TOTAL_SELF_SCORE_LABEL', 'Tổng điểm nhân sự tự đánh giá')
TOTAL_MAX_SCORE_LABEL = env('TOTAL_MAX_SCORE_LABEL', 'Tổng điểm tối đa')
RATING_LABEL = env('RATING_LABEL', 'Xếp loại tham khảo')

INFO_LABELS = env_list('INFO_LABELS')
MAIN_HEADERS = env_list('MAIN_HEADERS')
REFERENCE_SCALE_HEADERS = env_list('REFERENCE_SCALE_HEADERS')
REFERENCE_CRITERIA_HEADERS = env_list('REFERENCE_CRITERIA_HEADERS')
REFERENCE_RATING_HEADERS = env_list('REFERENCE_RATING_HEADERS')
CHECKLIST_HEADERS = env_list('CHECKLIST_HEADERS')
SCALE_ROWS = env_rows('SCALE_ROWS')
RATING_ROWS = env_rows('RATING_ROWS')
CHECKLIST_ITEMS = env_list('CHECKLIST_ITEMS')

navy = color('COLOR_NAVY', '0F172A')
blue = color('COLOR_BLUE', '2563EB')
purple = color('COLOR_PURPLE', '7C3AED')
cyan = color('COLOR_CYAN', 'DBEAFE')
light = color('COLOR_LIGHT', 'F8FAFC')
green = color('COLOR_GREEN', 'DCFCE7')
yellow = color('COLOR_YELLOW', 'FEF9C3')
red = color('COLOR_RED', 'FEE2E2')
orange = color('COLOR_ORANGE', 'FFEDD5')
white = color('COLOR_WHITE', 'FFFFFF')
guide = color('COLOR_GUIDE', 'EEF2FF')
muted_text = color('COLOR_MUTED_TEXT', '64748B')
dark_text = color('COLOR_DARK_TEXT', '334155')
gray_border = color('COLOR_GRAY_BORDER', 'CBD5E1')

thin_gray = Side(style='thin', color=gray_border)
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)


def style_cell(cell, fill=None, font=None, alignment=None):
    cell.border = border
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


def parse_sections():
    sections = []
    for item in env_list('SECTIONS'):
        if '@' not in item:
            continue
        title, merge_range = item.split('@', 1)
        first_cell = merge_range.split(':', 1)[0]
        title_row = int(''.join(ch for ch in first_cell if ch.isdigit())) - 1
        sections.append((f'A{title_row}', title.strip(), merge_range.strip()))
    return sections


def parse_rating_rules():
    parsed = []
    for group in env('RATING_RULES').split(';'):
        if ':' not in group:
            continue
        max_score, rules_text = group.split(':', 1)
        rules = []
        for rule in rules_text.split(','):
            if '=' not in rule:
                continue
            min_score, label = rule.split('=', 1)
            rules.append((int(min_score.strip()), label.strip()))
        parsed.append((int(max_score.strip()), rules))
    return parsed


def build_rating_formula(total_cell, max_cell):
    fallback = f'"{RATING_FALLBACK}"'
    formula = fallback
    for max_score, rules in reversed(parse_rating_rules()):
        inner = fallback
        for min_score, label in reversed(rules):
            inner = f'IF({total_cell}>={min_score},"{label}",{inner})'
        formula = f'IF({max_cell}={max_score},{inner},{formula})'
    return f'={formula}'


def read_criteria_from_env():
    criteria = []
    for row in env_rows('CRITERIA_ROWS'):
        if len(row) < 6:
            continue
        stt, name, weight, max_score, applies, desc = row[:6]
        criteria.append((int(stt), name, int(weight), desc, applies, int(max_score)))
    return criteria


criteria = read_criteria_from_env()
wb = Workbook()
ws = wb.active
ws.title = MAIN_SHEET

ws.merge_cells('A1:K1')
ws['A1'] = REPORT_TITLE
style_cell(ws['A1'], fill=navy, font=Font(bold=True, size=16, color=white), alignment=Alignment(horizontal='center', vertical='center'))
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:K2')
ws['A2'] = REPORT_SUBTITLE
style_cell(ws['A2'], fill=purple, font=Font(italic=True, size=11, color=white), alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))
ws.row_dimensions[2].height = 26

info_cells = [('A4', 'B4'), ('D4', 'E4'), ('G4', 'H4'), ('A5', 'B5'), ('D5', 'E5'), ('G5', 'H5')]
for index, (label_cell, value_cell) in enumerate(info_cells):
    label = INFO_LABELS[index] if index < len(INFO_LABELS) else ''
    ws[label_cell] = label
    style_cell(ws[label_cell], fill=cyan, font=Font(bold=True, color=navy), alignment=Alignment(vertical='center'))
    style_cell(ws[value_cell], fill=light, alignment=Alignment(vertical='center'))

ws.merge_cells('A7:K7')
ws['A7'] = f'Hướng dẫn: {GUIDE_TEXT}'
style_cell(ws['A7'], fill=guide, font=Font(italic=True, color=dark_text), alignment=Alignment(wrap_text=True, vertical='center'))
ws.row_dimensions[7].height = 32

start = 9
for col, header in enumerate(MAIN_HEADERS, 1):
    cell = ws.cell(start, col, header)
    style_cell(cell, fill=blue, font=Font(bold=True, color=white), alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))
ws.row_dimensions[start].height = 42

for r, item in enumerate(criteria, start + 1):
    stt, name, weight, desc, applies, max_score = item
    values = [stt, name, applies, weight, None, f'=D{r}*E{r}', max_score, '', desc, '', '']
    row_fill = white if stt <= 8 else ('F0FDF4' if stt == 9 else orange)
    for c, val in enumerate(values, 1):
        cell = ws.cell(r, c, val)
        align = Alignment(vertical='top', wrap_text=True)
        if c in [1, 4, 5, 6, 7, 10]:
            align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        style_cell(cell, fill=row_fill, alignment=align)
        if c in [4, 6, 7]:
            cell.font = Font(bold=True, color=navy)
        if c in [8, 9, 11]:
            cell.number_format = '@'

score_validation = DataValidation(type='whole', operator='between', formula1='1', formula2='5', allow_blank=True)
score_validation.error = 'Chỉ nhập điểm nguyên từ 1 đến 5'
score_validation.errorTitle = 'Điểm không hợp lệ'
ws.add_data_validation(score_validation)
score_validation.add(f'E{start+1}:E{start+len(criteria)}')
score_validation.add(f'J{start+1}:J{start+len(criteria)}')

summary_row = start + len(criteria) + 2
summary_items = [
    (f'C{summary_row}', TOTAL_SELF_SCORE_LABEL, cyan),
    (f'F{summary_row}', f'=SUM(F{start+1}:F{start+len(criteria)})', cyan),
    (f'G{summary_row}', f'=SUM(G{start+1}:G{start+len(criteria)})', cyan),
    (f'H{summary_row}', RATING_LABEL, green),
]
for cell_id, value, fill in summary_items:
    ws[cell_id] = value
    style_cell(ws[cell_id], fill=fill, font=Font(bold=True, color=navy), alignment=Alignment(horizontal='center', vertical='center'))
ws[f'I{summary_row}'] = build_rating_formula(f'F{summary_row}', f'G{summary_row}')
style_cell(ws[f'I{summary_row}'], fill=green, font=Font(bold=True, color=navy), alignment=Alignment(horizontal='center', vertical='center'))
ws.row_dimensions[summary_row].height = 28

for title_cell, title, merge_range in parse_sections():
    ws[title_cell] = title
    style_cell(ws[title_cell], fill=yellow, font=Font(bold=True, color=navy), alignment=Alignment(vertical='center'))
    ws[title_cell].number_format = '@'
    ws.merge_cells(merge_range)
    top_left = merge_range.split(':')[0]
    ws[top_left] = SECTION_PLACEHOLDER
    style_cell(ws[top_left], fill=light, font=Font(color=muted_text), alignment=Alignment(vertical='top', wrap_text=True))
    ws[top_left].number_format = '@'

ref = wb.create_sheet(REFERENCE_SHEET)
ref.merge_cells('A1:E1')
ref['A1'] = REFERENCE_TITLE
style_cell(ref['A1'], fill=navy, font=Font(bold=True, size=15, color=white), alignment=Alignment(horizontal='center', vertical='center'))
ref.row_dimensions[1].height = 32

ref.append([])
ref.append(REFERENCE_SCALE_HEADERS)
for row in SCALE_ROWS:
    ref.append(row)

ref.append([])
ref.append(REFERENCE_CRITERIA_HEADERS)
for stt, name, weight, _desc, _applies, max_score in criteria:
    ref.append([stt, name, weight, max_score])

ref.append([])
ref.append(REFERENCE_RATING_HEADERS)
for row in RATING_ROWS:
    ref.append(row)

ref.append([])
ref.append(CHECKLIST_HEADERS)
for item in CHECKLIST_ITEMS:
    ref.append([item, CHECKLIST_MARK])

for sheet in [ws, ref]:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    sheet.sheet_view.showGridLines = False

for row in ref.iter_rows():
    first_value = row[0].value
    if first_value in [REFERENCE_SCALE_HEADERS[0], REFERENCE_CRITERIA_HEADERS[0], CHECKLIST_HEADERS[0]]:
        for cell in row:
            if cell.value is not None:
                style_cell(cell, fill=blue, font=Font(bold=True, color=white), alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))

ws.freeze_panes = 'A10'
ws.auto_filter.ref = f'A{start}:K{start + len(criteria)}'
ref.freeze_panes = 'A3'

widths = {'A': 12, 'B': 36, 'C': 24, 'D': 12, 'E': 16, 'F': 16, 'G': 14, 'H': 40, 'I': 40, 'J': 12, 'K': 26}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for i, width in enumerate([10, 24, 72, 18, 18], 1):
    ref.column_dimensions[get_column_letter(i)].width = width
for r in range(10, 21):
    ws.row_dimensions[r].height = 62
for r in [25, 30, 36, 40, 45]:
    ws.row_dimensions[r].height = 64

for rng in [f'E{start+1}:E{start+len(criteria)}', f'J{start+1}:J{start+len(criteria)}']:
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=['2'], fill=PatternFill('solid', fgColor=red)))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['3'], fill=PatternFill('solid', fgColor=yellow)))
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['4'], fill=PatternFill('solid', fgColor=green)))

wb.save(OUTPUT)
print(OUTPUT)
